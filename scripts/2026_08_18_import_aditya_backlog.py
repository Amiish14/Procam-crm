"""
Import the pre-sales backlog Excel that the interns worked on.

The interns are no longer with the company, so we import the *data* but
skip the owner assignment — every lead lands in the admin's unassigned
bucket, ready to be re-routed to whoever picks up pre-sales next.

Sheets:
- 'New Projects'  — 55 rows, fully worked (Phone call / Intro mail /
   Meeting / RFQ / Remark / Follow-up all populated). Backfill with
   correct stage + activity dates.
- '11-05-26'      — 30 rows, freshly discovered leads. Enter as 'New'.

Idempotent — dedups on (company + project) case-insensitive. Also cleans
up any PRESALES01 / PRESALES02 login accounts a previous version of this
script may have seeded, and un-assigns any leads still pointing at them.

Run:
  cd /var/www/procam-crm
  env PYTHONPATH=. python scripts/2026_08_18_import_aditya_backlog.py \
      --file data/imports/aditya_backlog.xlsx
"""
import os
import sys
import re
import argparse
from datetime import datetime, date

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app import app, db, Lead, Employee, LeadStageHistory  # noqa

DEFAULT_FILE = 'data/imports/aditya_backlog.xlsx'

def cleanup_intern_users():
    """The pre-sales interns are no longer with the company — remove any
    login accounts a previous version of this script may have created,
    and null out the assigned_to / assigned_name fields on any lead that
    was tied to them so the leads live in the admin's unassigned bucket."""
    stale_codes = ['PRESALES01', 'PRESALES02']
    for code in stale_codes:
        emp = Employee.query.filter_by(emp_code=code).first()
        if emp:
            db.session.delete(emp)
            print(f'  Removed stale employee {code}')
    # Un-assign any leads still pointing at those codes
    n = (Lead.query
         .filter(Lead.assigned_to.in_(stale_codes))
         .update({Lead.assigned_to: None, Lead.assigned_name: None},
                 synchronize_session=False))
    if n:
        print(f'  Un-assigned {n} leads previously tied to stale interns')
    db.session.commit()


def _s(v):
    if v is None: return None
    s = str(v).strip()
    return s or None


def _num(v):
    if v in (None, ''): return None
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


DATE_PATTERNS = [
    '%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y', '%d-%m-%y',
    '%Y-%m-%d', '%m/%d/%Y', '%d.%m.%Y', '%d %b %Y',
]


def _first_date(v):
    """Extract the first date-like token from a value. The remark/follow-up
    columns often stack multiple dates on newlines — we want the *latest*
    (last) date the presales exec noted."""
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    if not s: return None
    # Find all date-like tokens; return the last one (most recent contact)
    tokens = re.findall(r'\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}', s)
    if not tokens: return None
    for tok in reversed(tokens):
        for fmt in DATE_PATTERNS:
            try:
                d = datetime.strptime(tok, fmt).date()
                if d.year < 100: d = d.replace(year=2000 + d.year)
                return d
            except ValueError:
                pass
    return None


def _parse_pic(cell):
    """Extract (name, designation, email, phone) from a
    'Mr.X - Designation, 9876..., x@y.com' contact string."""
    if not cell: return (None, None, None, None)
    s = str(cell).strip()
    # Remove LinkedIn URL tail
    s = re.sub(r'Linkedin\s*Id\s*:\s*https?://\S+\s*/?', '', s, flags=re.I).strip(' ,')
    # phone: run of digits + spaces, at least 10 chars long
    ph = None
    m = re.search(r'(?:\+?\d[\d\s\-]{8,}\d)', s)
    if m: ph = re.sub(r'\s+', ' ', m.group(0)).strip()
    # email
    em = None
    m = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', s)
    if m: em = m.group(0)
    # Strip phone + email from string, then name/designation is what remains
    core = s
    if ph: core = core.replace(ph, '')
    if em: core = core.replace(em, '')
    core = re.sub(r'\s+', ' ', core).strip(' ,')
    # Split "Mr.X - Designation" on the first ' - '
    name, desig = core, None
    if ' - ' in core:
        name, desig = core.split(' - ', 1)
    # Strip prefixes
    name = re.sub(r'^(Mr\.|Ms\.|Dr\.|Mrs\.|Shri\.?|Smt\.?)\s*', '', name).strip(' ,.')
    if desig: desig = desig.strip(' ,')
    return (name or None, desig or None, em, ph)


def _map_procam_pic(cell):
    """The pre-sales interns are no longer with the company, so we don't
    assign the imported leads to anyone. They land in the admin's
    unassigned bucket, ready to be re-routed later."""
    return (None, None)


def _classify_stage(phone_flag, mail_flag, meeting_val, rfq_val, remark):
    """Derive the new-flow stage from activity signals."""
    r = (str(remark or '')).lower()
    if rfq_val:
        return 'RFQ Generated'
    # Meeting done / scheduled → Visit Done or Appointment
    if meeting_val:
        # If the remark says "set up meeting" / "meeting is scheduled" → Appointment
        if 'set up' in r or 'schedul' in r or 'meeting on' in r or '@' in r:
            return 'Appointment'
        return 'Visit Done'
    # Mail sent → Profile Sent
    if str(mail_flag or '').strip().upper() in ('DONE', 'YES', 'SENT'):
        return 'Profile Sent'
    # Explicit rejection / not interested
    if any(k in r for k in ['not interested', 'rejected', 'not intersted']):
        return 'Not Interested'
    if any(k in r for k in ['on hold', 'later', 'call after', 'call some other']):
        return 'On Hold'
    # Phone call attempted (DONE) → Call Done
    if str(phone_flag or '').strip().upper() in ('DONE', 'YES'):
        return 'Call Done'
    return 'New'


def _build_history(company, phone_flag, mail_flag, mtg, rfq, remark, followup, procam_pic):
    """Compose a human-readable activity history similar to what the exec
    would have typed manually. Rendered under the lead detail 'History' box."""
    lines = []
    if phone_flag and str(phone_flag).strip().upper() == 'DONE':
        lines.append('• Phone call attempted')
    if mail_flag and str(mail_flag).strip().upper() == 'DONE':
        lines.append('• Intro mail sent')
    if mtg:
        lines.append(f'• Meeting: {mtg}')
    if rfq:
        lines.append(f'• RFQ received: {rfq}')
    if remark:
        rem = re.sub(r'\n+', ' | ', str(remark).strip())
        lines.append(f'• Remark: {rem}')
    if followup:
        fu = re.sub(r'\n+', ' | ', str(followup).strip())
        lines.append(f'• Follow-up scheduled: {fu}')
    # Owner names stripped — the interns who worked these leads have left.
    if not lines: return ''
    return '[Imported from pre-sales backlog on ' + \
           date.today().strftime('%d-%b-%Y') + ']\n' + '\n'.join(lines)


def _row_to_payload(r, headers_idx, sheet_name, imported_date):
    """Translate one Excel row into a Lead kwargs dict."""
    def col(name):
        ix = headers_idx.get(name)
        return r[ix] if ix is not None and ix < len(r) else None

    company = _s(col('Company'))
    if not company: return None
    project = _s(col('Project'))
    industry = _s(col('Industry'))
    cost = _num(col('Project Cost(Rs. Million)'))
    products = _s(col('Products & Capacity'))
    state = _s(col('Project State')) or _s(col('Addr. State'))
    city  = _s(col('City')) or _s(col('District'))
    address = _s(col('Address'))
    project_history = _s(col('Project History'))

    tel  = _s(col('Telephone'))
    mail = _s(col('Email'))
    pic  = _s(col('PIC'))
    p1   = _s(col('Person Name 1'))
    p2   = _s(col('Person Name 2'))

    # Primary contact = "Person Name 1" (that's usually the real decision maker
    # the exec noted). Fall back to the PIC column, then Person 2.
    primary_source = p1 or pic or p2
    name, desig, email1, phone1 = _parse_pic(primary_source)
    if not email1: email1 = mail
    if not phone1: phone1 = tel
    # Secondary contact = Person 2 (or PIC if we used P1 as primary)
    _, _, email2, phone2 = (None, None, None, None)
    if p2 and primary_source != p2:
        _, _, email2, phone2 = _parse_pic(p2)

    phone_flag = col('Phone call') or col('Phone call ')
    mail_flag  = col('Intro Mail Sent Date')
    mtg_val    = col('Meeting Date')
    rfq_val    = col('RFQ Recieved Date')
    remark     = col('Remark')
    followup   = col('Follow-up date')
    procam_pic = _s(col('Procam PIC'))

    stage = _classify_stage(phone_flag, mail_flag, mtg_val, rfq_val, remark)

    # Activity dates — the source columns are messy (contain 'DONE' or dates
    # or both). Fall back to the follow-up date as an anchor for the day
    # the exec was working the lead.
    followup_d = _first_date(followup)
    anchor = followup_d or imported_date
    phone_d = anchor if str(phone_flag or '').strip().upper() == 'DONE' else None
    mail_d  = anchor if str(mail_flag  or '').strip().upper() == 'DONE' else None
    mtg_d   = _first_date(mtg_val)  or (anchor if mtg_val else None)
    rfq_d   = _first_date(rfq_val)  or (anchor if rfq_val else None)

    emp_code, assigned_name = _map_procam_pic(procam_pic)

    # Build long-form notes: address + project history + first-line remark
    notes_bits = []
    if address: notes_bits.append(f'Address: {address}')
    if project_history: notes_bits.append(project_history)
    notes = '\n\n'.join(notes_bits) if notes_bits else None

    history = _build_history(company, phone_flag, mail_flag, mtg_val,
                              rfq_val, remark, followup, procam_pic)

    return dict(
        source='backlog_import',
        company=company[:200],
        project=(project or '')[:300] or None,
        industry=(industry or '')[:100] or None,
        cost_million=cost or 0,
        products=products,
        state=(state or '')[:60] or None,
        city=(city or '')[:60] or None,
        country='India',
        pic=(name or '')[:100] or None,
        designation_pic=(desig or '')[:100] or None,
        email=(email1 or '')[:120] or None,
        phone=(phone1 or '')[:30] or None,
        email2=(email2 or '')[:120] or None,
        phone2=(phone2 or '')[:30] or None,
        stage=stage,
        procam_vertical='Heavy Transport',
        assigned_to=emp_code,
        assigned_name=assigned_name,
        followup_date=followup_d,
        notes=notes,
        history=history,
        phone_call_date=phone_d,
        intro_mail_date=mail_d,
        meeting_date=mtg_d,
        rfq_date=rfq_d,
        onboarded_date=imported_date,
    )


def _iter_sheet(ws, has_header):
    """Yield row tuples + headers-index dict for a sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if has_header:
        # Header is on row 2 (idx 1) — row 1 is blank
        headers = [str(c or '').strip() for c in rows[1]]
        start = 2
    else:
        # 11-05-26: no header, same column order as New Projects
        headers = ['Sl No.', 'Company', 'Project', 'Project Type', 'Ownership',
                   'Industry', 'Project Cost(Rs. Million)', 'Products & Capacity',
                   'Completion Schedule', 'Project Stage', 'Location', 'District',
                   'Project State', 'Project History', 'Address', 'City',
                   'Pincode', 'Addr. State', 'Telephone', 'Email', 'PIC',
                   'Person Name 1', 'Person Name 2', 'Procam PIC',
                   'Phone call', 'Intro Mail Sent Date', 'Meeting Date',
                   'RFQ Recieved Date', 'Remark', 'Follow-up date', 'Lead stage']
        start = 0
    idx = {h: i for i, h in enumerate(headers)}
    for r in rows[start:]:
        if r and r[1]:   # Company col
            yield r, idx


def import_file(path, dry_run=False):
    if not os.path.isabs(path):
        path = os.path.join(os.getcwd(), path)
    if not os.path.exists(path):
        print(f'File not found: {path}'); sys.exit(1)

    with app.app_context():
        db.create_all()
        cleanup_intern_users()

        wb = openpyxl.load_workbook(path, data_only=True)
        today = date.today()

        n_new = n_upd = n_skip = 0
        for sheet_name in ['New Projects', '11-05-26']:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            has_header = (sheet_name == 'New Projects')
            print(f'\n=== {sheet_name} ===')
            for r, idx in _iter_sheet(ws, has_header):
                payload = _row_to_payload(r, idx, sheet_name, today)
                if not payload: n_skip += 1; continue

                # Dedup on (company + project) case-insensitive
                q = (Lead.query
                     .filter(db.func.lower(Lead.company) == payload['company'].lower()))
                if payload['project']:
                    q = q.filter((Lead.project == None)
                                 | (db.func.lower(Lead.project)
                                    == payload['project'].lower()))
                existing = q.first()

                if existing:
                    # Update only empty fields — don't clobber whatever the
                    # team has typed since. Stage is upgraded (never downgraded).
                    for k, v in payload.items():
                        if v is None or v == '' or v == 0: continue
                        cur = getattr(existing, k, None)
                        if not cur:
                            setattr(existing, k, v)
                    n_upd += 1
                    if not dry_run:
                        db.session.add(LeadStageHistory(
                            lead_id=existing.id, from_stage=existing.stage or '',
                            to_stage=payload['stage'],
                            note='Reconciled with pre-sales backlog Excel',
                            changed_by='backlog_import'))
                else:
                    lead = Lead(**payload)
                    db.session.add(lead)
                    db.session.flush()
                    db.session.add(LeadStageHistory(
                        lead_id=lead.id, from_stage='', to_stage=payload['stage'],
                        note='Imported from pre-sales backlog Excel',
                        changed_by='backlog_import'))
                    n_new += 1

                if (n_new + n_upd) % 20 == 0:
                    if not dry_run: db.session.commit()

        if dry_run:
            db.session.rollback()
            print('\n(dry run — rolled back)')
        else:
            db.session.commit()

        print(f'\n=== Done ===')
        print(f'  New leads    : {n_new}')
        print(f'  Updated leads: {n_upd}')
        print(f'  Skipped rows : {n_skip}')
        print(f'  Total leads  : {Lead.query.count()}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=DEFAULT_FILE)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    import_file(args.file, dry_run=args.dry_run)


if __name__ == '__main__':
    main()
